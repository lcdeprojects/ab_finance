import json
from decimal import Decimal

from dateutil.relativedelta import relativedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .forms import ClienteForm, PaymentForm
from .models import Cliente, Installment, Payment


@login_required
def dashboard(request):
    today = timezone.now().date()
    current_month = today.month
    current_year = today.year

    # Monthly Inflows: Sum of paid installments/entries for the current month based on PAYMENT DATE
    monthly_inflows = Installment.objects.filter(
        payment_date__month=current_month,
        payment_date__year=current_year,
        is_paid=True
    ).aggregate(total=Sum('paid_value'))['total'] or 0

    monthly_installments = Installment.objects.filter(
        due_date__month=current_month,
        due_date__year=current_year,
        is_paid=False
    )
    # Pending = Total Installment Value - Paid Value so far
    monthly_pending = sum(inst.pending_value for inst in monthly_installments)


    annual_total = Installment.objects.filter(
        payment_date__year=current_year,
        is_paid=True
    ).aggregate(total=Sum('paid_value'))['total'] or 0

    # Inadimplência: due_date < today and is_paid=False
    defaults_agg = Installment.objects.filter(
        due_date__lt=today,
        is_paid=False
    ).aggregate(v=Sum('value'), p=Sum('paid_value'))
    total_defaults = (defaults_agg['v'] or 0) - (defaults_agg['p'] or 0)

    # Total Received: Sum of all paid_value from all installments
    total_received = Installment.objects.filter(is_paid=True).aggregate(total=Sum('paid_value'))['total'] or 0
    
    # Total Outstanding: Sum of (value - paid_value) for all UNPAID installments
    pending_agg = Installment.objects.filter(is_paid=False).aggregate(v=Sum('value'), p=Sum('paid_value'))
    total_pending = (pending_agg['v'] or 0) - (pending_agg['p'] or 0)

    # Total in Contracts: Sum of all Payment total_value
    total_contracts = Payment.objects.aggregate(total=Sum('total_value'))['total'] or 0

    # Monthly Chart Data (Last 12 months)
    chart_labels = []
    chart_values = []
    
    MONTH_NAMES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    
    for i in range(11, -1, -1):
        target_date = today - relativedelta(months=i)
        m = target_date.month
        y = target_date.year
        
        m_inst = Installment.objects.filter(
            payment_date__month=m,
            payment_date__year=y,
            is_paid=True
        ).aggregate(total=Sum('paid_value'))['total'] or 0
        
        chart_labels.append(f"{MONTH_NAMES[m-1]}/{str(y)[2:]}")
        chart_values.append(float(m_inst))

    context = {
        'monthly_inflows': monthly_inflows,
        'monthly_pending': monthly_pending,
        'annual_total': annual_total,
        'total_defaults': total_defaults,
        'total_pending': total_pending,
        'total_historical': total_received,
        'total_contracts': total_contracts,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
    }
    return render(request, 'core/dashboard.html', context)

@login_required
def payment_create(request):
    cliente_id = request.GET.get('cliente_id')
    initial_data = {}
    if cliente_id:
        initial_data['cliente'] = get_object_or_404(Cliente, pk=cliente_id)

    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            messages.success(request, 'Pagamento cadastrado com sucesso! Parcelas geradas.')
            return redirect('payment_detail', pk=payment.pk)
    else:
        form = PaymentForm(initial=initial_data)
    
    return render(request, 'core/payment_form.html', {'form': form})

@login_required
def payment_detail(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    all_installments = payment.installment_set.all().order_by('number')
    
    # Separate entry (0) from normal installments (>0)
    entry = all_installments.filter(number=0).first()
    installments = all_installments.filter(number__gt=0)
    
    # Calculate statistics
    total_installments = installments.count()
    paid_installments = installments.filter(is_paid=True).count()
    open_installments = total_installments - paid_installments
    
    paid_value = all_installments.filter(is_paid=True).aggregate(total=Sum('paid_value'))['total'] or Decimal('0')
    
    today = timezone.now().date()
    val_open = sum(inst.pending_value for inst in all_installments.filter(is_paid=False, due_date__gte=today))
    val_default = sum(inst.pending_value for inst in all_installments.filter(is_paid=False, due_date__lt=today))
    
    context = {
        'payment': payment,
        'entry': entry,
        'installments': installments,
        'stats': {
            'total_count': total_installments,
            'paid_count': paid_installments,
            'open_count': open_installments,
            'paid_value': paid_value,
            'open_value': val_open,
            'default_value': val_default,
            'total_value': payment.total_value
        }
    }
    return render(request, 'core/payment_detail.html', context)

@login_required
def cliente_list(request):
    name_query = request.GET.get('name')
    qs = Cliente.objects.all()
    if name_query:
        qs = qs.filter(nome__icontains=name_query)
    
    clientes = qs.order_by('nome')
    return render(request, 'core/client_list.html', {'clientes': clientes, 'name_query': name_query})

@login_required
def cliente_create(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            messages.success(request, 'Cliente cadastrado com sucesso!')
            return redirect('cliente_detail', pk=cliente.pk)
    else:
        form = ClienteForm()
    return render(request, 'core/client_form.html', {'form': form})

@login_required
def cliente_update(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cadastro do cliente atualizado!')
            return redirect('cliente_list')
    else:
        form = ClienteForm(instance=cliente)
    return render(request, 'core/client_form.html', {'form': form, 'is_update': True})

@login_required
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        from django.contrib import messages
        messages.success(request, 'Cliente excluído com sucesso!')
        return redirect('cliente_list')
    return render(request, 'core/cliente_confirm_delete.html', {'cliente': cliente})

@login_required
def cliente_detail(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    payments = cliente.payments.all().order_by('-created_at')
    total_invested = payments.aggregate(total=Sum('total_value'))['total'] or 0
    
    context = {
        'cliente': cliente,
        'payments': payments,
        'total_invested': total_invested
    }
    return render(request, 'core/client_detail.html', context)

# Helper to get month string
MONTHS = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
}

@login_required
def payment_control(request):
    month_val = request.GET.get('month')
    year_val = request.GET.get('year')
    name_query = request.GET.get('name')
    status_filter = request.GET.get('status')
    
    today = timezone.now().date()
    
    # Default to current month/year if not provided
    month = int(month_val) if month_val else today.month
    year = int(year_val) if year_val else today.year
    
    # Base queryset: only UNPAID installments
    qs = Installment.objects.filter()
    
    if month and year:
        qs = qs.filter(due_date__month=month, due_date__year=year)
    
    if name_query:
        qs = qs.filter(payment__cliente__nome__icontains=name_query)
        
    if status_filter == 'overdue':
        qs = qs.filter(due_date__lt=today, is_paid=False)
    if status_filter == 'scheduled':
        qs = qs.filter(due_date__gte=today, is_paid=False)
    if status_filter == 'paid':
        qs = qs.filter(is_paid=True)
        
    installments = qs.order_by('due_date')
    
    total_to_receive = sum(inst.pending_value for inst in installments)

    context = {
        'installments': installments,
        'selected_month': month,
        'selected_year': year,
        'month_name': MONTHS.get(month, "Mês"),
        'months': MONTHS,
        'total_to_receive': total_to_receive,
        'name_query': name_query,
        'status_filter': status_filter,
        'today': timezone.now().date(),
    }
    return render(request, 'core/control_list.html', context)

@login_required
def default_list(request):
    month_val = request.GET.get('month')
    year_val = request.GET.get('year')
    name_query = request.GET.get('name')
    
    today = timezone.now().date()
    
    # Base queryset: only UNPAID and OVERDUE (due_date < today)
    qs = Installment.objects.filter(due_date__lt=today, is_paid=False)
    
    if month_val:
        qs = qs.filter(due_date__month=month_val)
    if year_val:
        qs = qs.filter(due_date__year=year_val)
    if name_query:
        qs = qs.filter(payment__cliente__nome__icontains=name_query)
        
    defaults = qs.order_by('due_date')
    total_defaults = sum(inst.pending_value for inst in defaults)
    
    context = {
        'defaults': defaults,
        'total_defaults': total_defaults,
        'selected_month': int(month_val) if month_val else None,
        'selected_year': int(year_val) if year_val else None,
        'name_query': name_query,
        'months': MONTHS,
        'today': today,
    }
    return render(request, 'core/default_list.html', context)

@login_required
def installment_update(request, pk):
    installment = get_object_or_404(Installment, pk=pk)
    if request.method == 'POST':
        paid_value_str = request.POST.get('paid_value', '0').replace(',', '.')
        try:
            paid_value = Decimal(paid_value_str)
        except:
            paid_value = Decimal('0')
            
        is_paid = request.POST.get('is_paid') == 'on'
        has_nf = request.POST.get('has_nf') == 'on'
        payment_date_str = request.POST.get('payment_date')
        
        payment_date = None
        if is_paid:
            if payment_date_str:
                try:
                    payment_date = payment_date_str
                except:
                    payment_date = timezone.now().date()
            else:
                payment_date = timezone.now().date()
        
        # If user checked "Paid" but left value as 0, assume total payment
        if is_paid and paid_value == 0:
            paid_value = installment.value
            
        # Logic for partial payment split
        if is_paid and paid_value < installment.value and paid_value > 0:
            remaining_value = installment.value - paid_value
            # Update current as paid with the partial value
            installment.value = paid_value
            installment.paid_value = paid_value
            installment.is_paid = True
            installment.has_nf = has_nf
            installment.payment_date = payment_date
            installment.save()
            
            # Create new installment for the remaining value
            Installment.objects.create(
                payment=installment.payment,
                number=installment.number,
                value=remaining_value,
                due_date=installment.due_date,
                is_paid=False,
                has_nf=False
            )
            messages.info(request, f"Pagamento parcial de R$ {paid_value} realizado. Saldo de R$ {remaining_value} em aberto.")
        elif not is_paid:
            # If unchecked, reset paid value
            installment.paid_value = 0
            installment.is_paid = False
            installment.has_nf = has_nf
            installment.payment_date = None
            installment.save()
        else:
            # Full payment or update
            installment.paid_value = paid_value
            installment.is_paid = is_paid
            installment.has_nf = has_nf
            installment.payment_date = payment_date
            installment.save()
            if is_paid:
                messages.success(request, f"Pagamento de R$ {paid_value} confirmado!")
        
        # If HTMX request and payment is confirmed, refresh the full list 
        # to ensure the row disappears from "Controle Mensal"
        response = render(request, 'core/partials/installment_row.html', {'inst': installment})
        if is_paid:
            response['HX-Refresh'] = 'true'
        return response

@login_required
def paid_installment_delete(request, pk):
    installment = get_object_or_404(Installment, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'paid_list'
    
    if request.method == 'POST':
        installment.is_paid = False
        installment.paid_value = 0
        installment.save()
        messages.success(request, 'Pagamento removido do histórico.')
        return redirect(next_url)
    return render(request, 'core/payment_confirm_delete.html', {
        'installment': installment, 
        'is_installment': True,
        'next_url': next_url
    })

@login_required
def payment_update(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            # In a full app, checking if values changed to recalculate installments would be here
            form.save()
            messages.success(request, 'Pagamento atualizado com sucesso!')
            return redirect('dashboard')
    else:
        form = PaymentForm(instance=payment)
    return render(request, 'core/payment_form.html', {'form': form, 'is_update': True, 'payment': payment})

@login_required
def payment_delete(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    next_url = request.GET.get('next') or request.POST.get('next') or 'payment_list'
    
    if request.method == 'POST':
        payment.delete() # Cascades to installments
        messages.success(request, 'Pagamento e suas parcelas foram excluídos!')
        return redirect(next_url)
    return render(request, 'core/payment_confirm_delete.html', {
        'payment': payment,
        'next_url': next_url
    })

@login_required
def payment_list(request):
    payments = Payment.objects.all().order_by('-created_at')
    name_query = request.GET.get('name')
    if name_query:
        payments = payments.filter(cliente__nome__icontains=name_query)
    return render(request, 'core/payment_list.html', {'payments': payments})

@login_required
def paid_list(request):
    month_val = request.GET.get('month')
    year_val = request.GET.get('year')
    name_query = request.GET.get('name')
    
    today = timezone.now().date()
    
    # Default to current month/year if not provided for receipts
    month = int(month_val) if month_val else today.month
    year = int(year_val) if year_val else today.year
    
    # Base queryset: only PAID installments filtered by payment date
    qs = Installment.objects.filter(payment_date__month=month, payment_date__year=year, is_paid=True)
    
    if name_query:
        qs = qs.filter(payment__cliente__nome__icontains=name_query)
        
    paid_installments = qs.order_by('-due_date')
    total_received = sum(inst.paid_value for inst in paid_installments)

    context = {
        'installments': paid_installments,
        'selected_month': month,
        'selected_year': year,
        'month_name': MONTHS.get(month, "Mês"),
        'months': MONTHS,
        'total_paid': total_received,
        'name_query': name_query,
    }
    return render(request, 'core/paid_list.html', context)

@login_required
def pay_list(request):
    month_val = request.GET.get('month')
    year_val = request.GET.get('year')
    name_query = request.GET.get('name')
    
    today = timezone.now().date()
    
    # Default to current month/year if not provided for receipts
    month = int(month_val) if month_val else today.month
    year = int(year_val) if year_val else today.year
    
    # Base queryset: only PAID installments filtered by payment date
    qs = Installment.objects.filter(payment_date__month=month, payment_date__year=year)
    
    if name_query:
        qs = qs.filter(payment__cliente__nome__icontains=name_query)
        
    paid_installments = qs.order_by('-due_date')
    total_received = sum(inst.value for inst in paid_installments)

    context = {
        'installments': paid_installments,
        'selected_month': month,
        'selected_year': year,
        'month_name': MONTHS.get(month, "Mês"),
        'months': MONTHS,
        'month_total': total_received,
        'name_query': name_query,
    }
    return render(request, 'core/pay_list.html', context)

@login_required
def export_installments_excel(request):
    source = request.GET.get('source', 'control')
    month_val = request.GET.get('month')
    year_val = request.GET.get('year')
    name_query = request.GET.get('name')
    
    today = timezone.now().date()
    
    # Filter Logic (Identical to views but unified)
    if source == 'paid':
        month = int(month_val) if month_val else today.month
        year = int(year_val) if year_val else today.year
        qs = Installment.objects.filter(due_date__month=month, due_date__year=year, is_paid=True)
        filename = f"recebidos_{month}_{year}.xlsx"
    elif source == 'defaults':
        qs = Installment.objects.filter(due_date__lt=today, is_paid=False)
        if month_val: qs = qs.filter(due_date__month=month_val)
        if year_val: qs = qs.filter(due_date__year=year_val)
        filename = "inadimplencia.xlsx"
    else: # control
        month = int(month_val) if month_val else today.month
        year = int(year_val) if year_val else today.year
        qs = Installment.objects.filter(due_date__month=month, due_date__year=year)
        filename = f"controle_mensal_{month}_{year}.xlsx"

    if name_query:
        qs = qs.filter(payment__cliente__nome__icontains=name_query)

    installments = qs.order_by('due_date', 'payment__cliente__nome')

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Financeiro"

    # Headers
    headers = ['Cliente', 'Parcela', 'Método', 'Valor Total', 'Vencimento', 'Data Pagamento', 'Status', 'Valor Pago', 'NF']
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for inst in installments:
        status = "Pago" if inst.is_paid else ("Inadimplente" if inst.due_date < today else "A Vencer")
        ws.append([
            inst.payment.cliente.nome,
            f"{inst.number}/{inst.payment.installments}",
            inst.payment.payment_method,
            float(inst.value),
            inst.due_date.strftime('%d/%m/%Y'),
            inst.payment_date.strftime('%d/%m/%Y') if inst.payment_date else '-',
            status,
            float(inst.paid_value),
            "Sim" if inst.has_nf else "Não"
        ])

    # Auto-adjust columns width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 5

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
